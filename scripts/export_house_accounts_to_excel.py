#!/usr/bin/env python3
"""Export Shopify customers matching a tag filter to an Excel file,
classify each as Personal vs Business vs Unclear, and open in Excel on macOS.

Defaults to tag:house-account -> /Users/moosemarketer/Code/garoppos/tmp/house_accounts.xlsx.
All parameters are constants at the top — edit for reuse.
"""

import argparse
import json
import logging
import os
import re
import subprocess
import sys
import time

import requests

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from uploader_modules.config import load_config

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

API_VERSION = "2025-10"

# =============================================================================
# PARAMETERS — edit here for reuse
# =============================================================================
SOURCE_TAG   = "house-account"
OUTPUT_PATH  = "/Users/moosemarketer/Code/garoppos/tmp/house_accounts.xlsx"
SHEET_NAME   = "House Accounts"

# Classification keywords (case-insensitive). First match wins.
BUSINESS_KEYWORDS = [
    "LLC", "Inc", "Inc.", "Corp", "Corporation", "Co.", "Company",
    "Construction", "Contractor", "Contracting",
    "Landscaping", "Landscape", "Services", "Supply", "Enterprises",
    "Group", "Holdings", "Properties", "Nursery", "Garden Center",
    "Farm", "Farms", "Memorial", "Homes", "Builders", "Masonry",
    "Excavating", "Roofing", "Plumbing", "Electric", "Fence",
    "Pools", "Pool", "Pavers", "Hardscape", "Utility",
    "PSE&G", "PSEG", "Township", "Borough", "City of", "County",
    "Church", "School", "LLP", "LP", "PA", "PLLC",
]

# Match against word boundaries so "Inc" doesn't match "Incandescent".
# Some keywords contain special chars — handle those by escaping.
def build_keyword_patterns():
    patterns = []
    for kw in BUSINESS_KEYWORDS:
        esc = re.escape(kw)
        # Use word boundaries for alphabetic-start/end keywords; looser for symbolic ones
        if re.search(r"[A-Za-z]", kw[0]) and re.search(r"[A-Za-z]", kw[-1]):
            patterns.append((kw, re.compile(rf"\b{esc}\b", re.IGNORECASE)))
        else:
            patterns.append((kw, re.compile(esc, re.IGNORECASE)))
    return patterns

KEYWORD_PATTERNS = build_keyword_patterns()
HUMAN_NAME_RE = re.compile(r"^[A-Z][a-z]+(?:[- ][A-Z]?[a-z]+)?\s+[A-Z][a-z]+(?:[- ][A-Z]?[a-z]+)?\.?$")

# =============================================================================
# Shopify API
# =============================================================================
def api_url(cfg):
    s = cfg["SHOPIFY_STORE_URL"].replace("https://", "").replace("http://", "").strip("/")
    return f"https://{s}/admin/api/{API_VERSION}/graphql.json"

def headers(cfg):
    return {"Content-Type": "application/json",
            "X-Shopify-Access-Token": cfg["SHOPIFY_ACCESS_TOKEN"]}

def gql(cfg, q, v=None, retries=5):
    for _ in range(retries):
        r = requests.post(api_url(cfg), json={"query": q, "variables": v or {}},
                          headers=headers(cfg), timeout=60)
        r.raise_for_status()
        d = r.json()
        if "errors" in d:
            if any("THROTTLED" in str(e).upper() for e in d["errors"]):
                time.sleep(2); continue
            raise RuntimeError(f"GraphQL errors: {d['errors']}")
        return d["data"]
    raise RuntimeError("retries exhausted")

CUSTOMERS_Q = """
query($q: String!, $cursor: String) {
  customers(first: 100, query: $q, after: $cursor) {
    pageInfo { hasNextPage endCursor }
    edges { node {
      id displayName firstName lastName email phone tags
      createdAt updatedAt
      numberOfOrders
      amountSpent { amount currencyCode }
      defaultAddress { city province country address1 }
    } }
  }
}
"""

def fetch_customers(cfg, tag):
    customers = []
    cursor = None
    page = 0
    while True:
        page += 1
        d = gql(cfg, CUSTOMERS_Q, {"q": f"tag:{tag}", "cursor": cursor})
        edges = d["customers"]["edges"]
        for e in edges:
            customers.append(e["node"])
        pi = d["customers"]["pageInfo"]
        log.info(f"  page {page}: fetched {len(edges)}, total {len(customers)}, hasNext={pi['hasNextPage']}")
        if not pi["hasNextPage"]:
            break
        cursor = pi["endCursor"]
        time.sleep(0.15)
    return customers

# =============================================================================
# Classification
# =============================================================================
def classify(customer):
    """Return (account_type, reason)."""
    name = customer.get("displayName") or ""
    tags = [t.lower() for t in (customer.get("tags") or [])]

    # Rule 1: explicit business tag
    if "business" in tags:
        return "Business", "business tag"

    # Rule 2: keyword match
    for kw, pat in KEYWORD_PATTERNS:
        if pat.search(name):
            return "Business", f"keyword: {kw}"

    # Rule 3: looks like a human name (First Last, maybe with title-case)
    if HUMAN_NAME_RE.match(name.strip()):
        return "Personal", "human name pattern"

    return "Unclear", "ambiguous — no strong signal"

# =============================================================================
# Excel output
# =============================================================================
def write_excel(rows, output_path, sheet_name=SHEET_NAME):
    # Lazy import so missing openpyxl gives a clear error
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl is not installed. Run: pip install openpyxl")
        raise

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    columns = [
        ("Display Name",       "displayName",          32),
        ("Account Type",       "accountType",          14),
        ("Classification Reason","classificationReason", 28),
        ("Tags",               "tagsJoined",           42),
        ("Email",              "email",                26),
        ("Phone",              "phone",                16),
        ("City",               "city",                 16),
        ("State",              "province",             8),
        ("Orders",             "numberOfOrders",       8),
        ("Lifetime Spend",     "amountSpent",          14),
        ("Currency",           "currency",             10),
        ("Created",            "createdAt",            20),
        ("Updated",            "updatedAt",            20),
        ("Customer GID",       "id",                   44),
    ]
    # Header row
    for col_idx, (hdr, _, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (_, key, _) in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(key))

    # Column widths
    for c_idx, (_, _, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    # Freeze header, auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Conditional coloring on "Account Type" column (column B)
    fills = {
        "Business": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # green
        "Personal": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # blue
        "Unclear":  PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # yellow
    }
    for r_idx in range(2, len(rows) + 2):
        cell = ws.cell(row=r_idx, column=2)  # column B
        fill = fills.get(cell.value)
        if fill:
            cell.fill = fill

    wb.save(output_path)
    log.info(f"Wrote {len(rows)} rows to {output_path}")
    return output_path

# =============================================================================
# Main
# =============================================================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tag", default=SOURCE_TAG,
                    help=f"Customer tag filter (default: {SOURCE_TAG!r})")
    ap.add_argument("--output", default=OUTPUT_PATH,
                    help=f"Output .xlsx path (default: {OUTPUT_PATH!r})")
    ap.add_argument("--no-open", action="store_true", help="Skip opening Excel")
    args = ap.parse_args()

    cfg = load_config()

    # Fetch
    log.info(f"Querying customers with tag={args.tag!r}...")
    customers = fetch_customers(cfg, args.tag)
    log.info(f"Total: {len(customers)}")

    # Shape + classify
    rows = []
    for c in customers:
        name = c.get("displayName") or ""
        addr = c.get("defaultAddress") or {}
        amount = c.get("amountSpent") or {}
        account_type, reason = classify(c)
        rows.append({
            "id": c["id"],
            "displayName": name,
            "email": c.get("email") or "",
            "phone": c.get("phone") or "",
            "tagsJoined": ", ".join(c.get("tags") or []),
            "city": addr.get("city") or "",
            "province": addr.get("province") or "",
            "numberOfOrders": c.get("numberOfOrders"),
            "amountSpent": amount.get("amount") or "",
            "currency": amount.get("currencyCode") or "",
            "createdAt": c.get("createdAt") or "",
            "updatedAt": c.get("updatedAt") or "",
            "accountType": account_type,
            "classificationReason": reason,
        })

    # Sort alphabetically by display name (case-insensitive)
    rows.sort(key=lambda r: (r["displayName"] or "").lower())

    # Breakdown
    from collections import Counter
    breakdown = Counter(r["accountType"] for r in rows)
    log.info(f"Breakdown: {dict(breakdown)}")

    # Write
    write_excel(rows, args.output)

    # Report Unclear rows
    unclear = [r for r in rows if r["accountType"] == "Unclear"]
    if unclear:
        log.info(f"Unclear customers needing human review ({len(unclear)}):")
        for r in unclear:
            log.info(f"  - {r['displayName']!r}  tags=[{r['tagsJoined']}]  orders={r['numberOfOrders']}  id={r['id']}")

    # Open in Excel
    if not args.no_open:
        log.info(f"Opening {args.output} in Excel...")
        try:
            subprocess.run(["open", "-a", "Microsoft Excel", args.output], check=True)
            log.info("  Excel open issued")
        except subprocess.CalledProcessError:
            log.warning("Microsoft Excel not available, falling back to default handler")
            subprocess.run(["open", args.output], check=False)

    print()
    print("=" * 70)
    print("REPORT")
    print("=" * 70)
    print(f"File:             {args.output}")
    print(f"Rows:             {len(rows)}")
    print(f"Business:         {breakdown.get('Business', 0)}")
    print(f"Personal:         {breakdown.get('Personal', 0)}")
    print(f"Unclear:          {breakdown.get('Unclear', 0)}")
    if unclear:
        print(f"Unclear entries for review:")
        for r in unclear:
            print(f"  - {r['displayName']!r}  tags=[{r['tagsJoined']}]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
